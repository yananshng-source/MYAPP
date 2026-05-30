# app.py
import streamlit as st
import os
import pandas as pd
from io import BytesIO
from PIL import Image, ImageOps
import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import re
from datetime import datetime
import logging
import traceback
from typing import Iterable, Any
import numpy as np
import subprocess
import sys
import tempfile
import zipfile
from PIL import ImageEnhance
from openpyxl.utils import get_column_letter
import uuid

# ------------------------ Config ------------------------
st.set_page_config(page_title="综合处理工具箱", layout="wide")
DEFAULT_TIMEOUT = 15
REQUEST_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
VERIFY_SSL = False  # cloud 上有些站点会证书问题，保守设为 False
MAX_LOG_LINES = 200

# ------------------------ Logging ------------------------
logger = logging.getLogger("综合处理工具箱")
logger.setLevel(logging.INFO)
if not logger.handlers:
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logger.addHandler(ch)

# store recent logs in session state to show in UI
if "recent_logs" not in st.session_state:
    st.session_state.recent_logs = []


def log(msg, level="info"):
    entry = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {level.upper()} - {msg}"
    st.session_state.recent_logs.append(entry)
    # cap length
    if len(st.session_state.recent_logs) > MAX_LOG_LINES:
        st.session_state.recent_logs = st.session_state.recent_logs[-MAX_LOG_LINES:]
    if level == "info":
        logger.info(msg)
    elif level == "warning":
        logger.warning(msg)
    elif level == "error":
        logger.error(msg)
    else:
        logger.debug(msg)


# ------------------------ Helpers ------------------------
def progress_iter(it: Iterable[Any], text="处理中...", progress_key=None):
    """
    Generic iterator wrapper that updates a single st.progress bar (main bar).
    It expects an iterable with a determinable length (like list, tuple, DataFrame rows via list()).
    Yields the original items.
    """
    # normalize to list to calculate total reliably (this will hold items in memory)
    # For very large iterables you may replace with a custom strategy.
    items = list(it)
    total = len(items)
    if progress_key is None:
        progress_key = "main_progress"
    progress_bar = st.session_state.get(progress_key)
    if progress_bar is None:
        progress_bar = st.progress(0, text=text)
        st.session_state[progress_key] = progress_bar
    try:
        for idx, item in enumerate(items):
            yield item
            percent = int((idx + 1) / total * 100) if total > 0 else 100
            try:
                progress_bar.progress(percent, text=text)
            except Exception:
                # fallback: ignore progress update error
                pass
        try:
            progress_bar.progress(100, text=text + " ✅ 完成")
        except Exception:
            pass
    finally:
        # clear stored progress bar so future calls get a fresh widget
        if progress_key in st.session_state:
            del st.session_state[progress_key]


def safe_requests_get(session: requests.Session, url: str, **kwargs):
    """
    Wrapper around session.get with global headers, timeout, verify options and robust exception handling.
    Returns response or raises.
    """
    try:
        resp = session.get(url, timeout=kwargs.get("timeout", DEFAULT_TIMEOUT),
                           headers=REQUEST_HEADERS, verify=VERIFY_SSL)
        resp.raise_for_status()
        return resp
    except Exception as e:
        raise


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path


def create_zip_download(files, zip_name="downloaded_images.zip"):
    """创建ZIP文件供下载"""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file_path in files:
            if os.path.exists(file_path):
                zip_file.write(file_path, os.path.basename(file_path))
    zip_buffer.seek(0)
    return zip_buffer


# ------------------------ Core functions ------------------------
def fix_mojibake(text):
    """修复常见的乱码问题"""
    if not isinstance(text, str):
        return text

    # UTF-8字节被错误解码为Latin-1的常见情况
    fixes = {
        'ÃƒÂ©': 'é', 'ÃƒÂ¨': 'è', 'ÃƒÂª': 'ê', 'ÃƒÂ§': 'ç',
        'ÃƒÂ¹': 'ù', 'ÃƒÂ»': 'û', 'ÃƒÂ®': 'î', 'ÃƒÂ¯': 'ï',
        'ÃƒÂ´': 'ô', 'ÃƒÂ¶': 'ö', 'ÃƒÂ¼': 'ü', 'ÃƒÂ¤': 'ä',
        'ÃƒÂ¥': 'å', 'ÃƒÂ¦': 'æ', 'ÃƒÂ¸': 'ø', 'ÃƒÂ¿': 'ÿ',
        'Ã©': 'é', 'Ã¨': 'è', 'Ãª': 'ê', 'Ã§': 'ç',
        'Ã¹': 'ù', 'Ã»': 'û', 'Ã®': 'î', 'Ã¯': 'ï',
        'Ã´': 'ô', 'Ã¶': 'ö', 'Ã¼': 'ü', 'Ã¤': 'ä',
        'Ã¥': 'å', 'Ã¦': 'æ', 'Ã¸': 'ø', 'Ã¿': 'ÿ',
        'â€¢': '·', 'â€"': '—', 'â€¦': '…', 'â€˜': "'",
        'â€™': "'", 'â€œ': '"', 'â€': '"', 'â€”': '—',
        'â€"': '—', 'â€"': '—', 'â€"': '—',
        'Â': '', 'Â ': ' ', 'Â ': ' ',  # 移除多余的空白字符
        'å': '•', 'æ': '•', 'è': '·', 'é': '·',
        '¡¯': "'", '¡±': '"', '¡°': '"',
        'ï¼ˆ': '（', 'ï¼‰': '）', 'ï¼š': '：',
        'ï¼Œ': '，', 'ï¼': '！', 'ï¼Ÿ': '？',
        'ï¼›': '；', 'ï¼€': '￥'
    }

    for wrong, right in fixes.items():
        text = text.replace(wrong, right)

    return text


def clean_dataframe_encoding(df):
    """清理DataFrame中的编码问题"""
    df_clean = df.copy()

    for col in df_clean.columns:
        if df_clean[col].dtype == 'object':
            # 尝试清理字符串
            df_clean[col] = df_clean[col].apply(
                lambda x: fix_mojibake(x) if isinstance(x, str) else x
            )

    return df_clean


def scrape_table(url_list, fill_cols=None, progress_callback=None):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

        total = len(url_list)

        for i, url in enumerate(url_list):
            try:
                tables = smart_fetch(url)

                if not tables:
                    log(f"{url} 未抓到表", "warning")
                    continue

                # 👉 合并当前网页的所有表（仅当前页）
                df = pd.concat(tables, ignore_index=True, sort=False)

                # 👉 填充（可选）
                df, used_cols = smart_fill(df, fill_cols)

                # 👉 Sheet 名处理（防止超长/非法）
                sheet_name = f"Sheet_{i+1}"
                sheet_name = sheet_name[:31]

                df.to_excel(writer, index=False, sheet_name=sheet_name)

                log(f"{url} → 写入 {sheet_name}（{len(df)} 行）")

            except Exception as e:
                log(f"{url} 失败: {e}", "error")

            if progress_callback:
                progress_callback(i + 1, total)

    output.seek(0)
    return output

def download_images_from_urls(url_list, output_dir=None):
    """
    从每个页面抓取 <img> 并下载。
    返回 (output_dir, downloaded_file_paths, errors)
    """
    # 在云环境中使用临时目录
    if output_dir is None:
        # 尝试创建桌面目录，如果失败则使用临时目录
        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "downloaded_images")
            ensure_dir(desktop_path)
            # 测试写入权限
            test_file = os.path.join(desktop_path, "test_write.txt")
            with open(test_file, 'w') as f:
                f.write("test")
            os.remove(test_file)
            output_dir = desktop_path
        except (PermissionError, OSError):
            # 如果没有桌面写入权限，使用临时目录
            output_dir = os.path.join(tempfile.gettempdir(), "downloaded_images")
            ensure_dir(output_dir)

    log(f"📁 图片下载目录: {output_dir}")

    session = requests.Session()
    session.headers.update(REQUEST_HEADERS)
    downloaded_files = []
    errors = []

    enumerated = list(enumerate(url_list, start=1))
    for idx, url in progress_iter(enumerated, text="下载网页图片中"):
        try:
            _, page_url = (idx, url)
            log(f"正在访问: {page_url}")
            resp = safe_requests_get(session, page_url)
            soup = BeautifulSoup(resp.content, "html.parser")
            title_tag = soup.find("title")
            title = title_tag.string.strip() if title_tag and title_tag.string else f"网页{idx}"
            safe_title = "".join([c if c not in r'\/:*?"<>|' else "_" for c in title])

            imgs = soup.find_all("img")
            log(f"📄 {page_url} - 找到 {len(imgs)} 张图片")

            if not imgs:
                log(f"{page_url} - 未找到 img 标签", level="info")
                continue

            for i, img_tag in enumerate(imgs, start=1):
                src = img_tag.get("src") or img_tag.get("data-src") or img_tag.get("data-original")
                if not src:
                    continue

                full_url = urljoin(page_url, src.strip())
                log(f"正在下载图片: {full_url}")

                try:
                    resp_img = safe_requests_get(session, full_url)

                    # 文件扩展名处理
                    ext = os.path.splitext(full_url.split('?')[0])[1]
                    if not ext or len(ext) > 6:
                        content_type = resp_img.headers.get('content-type', '')
                        if 'jpeg' in content_type or 'jpg' in content_type:
                            ext = ".jpg"
                        elif 'png' in content_type:
                            ext = ".png"
                        elif 'gif' in content_type:
                            ext = ".gif"
                        else:
                            ext = ".jpg"

                    fname = f"{safe_title}_{i:02d}{ext}"
                    fpath = os.path.join(output_dir, fname)

                    # 避免文件名重复
                    counter = 1
                    original_fpath = fpath
                    while os.path.exists(fpath):
                        name_only = os.path.splitext(original_fpath)[0]
                        fpath = f"{name_only}_{counter}{ext}"
                        counter += 1

                    with open(fpath, "wb") as f:
                        f.write(resp_img.content)

                    downloaded_files.append(fpath)
                    log(f"✅ 下载成功: {os.path.basename(fpath)} - 大小: {len(resp_img.content)} bytes")

                except Exception as e:
                    error_msg = f"图片下载失败: {full_url} -> {repr(e)}"
                    errors.append(error_msg)
                    log(error_msg, level="warning")
                    continue

        except Exception as e:
            error_msg = f"页面请求失败: {url} -> {repr(e)}"
            log(error_msg, level="warning")
            errors.append(error_msg)
            continue

    log(f"🎉 下载完成! 总共下载 {len(downloaded_files)} 张图片到 {output_dir}")
    return output_dir, downloaded_files, errors


def crop_images_only(folder_path, x_center, y_center, crop_width, crop_height):
    output_folder = os.path.join(os.path.expanduser("~"), "Desktop", "crop_results")
    ensure_dir(output_folder)
    img_exts = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff")
    filenames = [f for f in os.listdir(folder_path) if f.lower().endswith(img_exts)]
    for filename in progress_iter(filenames, text="裁剪图片中"):
        try:
            image_path = os.path.join(folder_path, filename)
            img = Image.open(image_path).convert("RGB")
            width, height = img.size
            left = max(0, int(x_center - crop_width // 2))
            right = min(width, int(x_center + crop_width // 2))
            top = max(0, int(y_center - crop_height // 2))
            bottom = min(height, int(y_center + crop_height // 2))
            crop_img = img.crop((left, top, right, bottom))
            # 放大二倍用于后续识别/查看
            crop_img = crop_img.resize((crop_img.width * 2, crop_img.height * 2), Image.LANCZOS)
            bw = ImageOps.grayscale(crop_img)
            save_path = os.path.join(output_folder, f"crop_{filename}")
            bw.save(save_path)
            log(f"裁剪并保存: {save_path}")
        except Exception as e:
            log(f"裁剪失败: {filename} -> {e}", level="warning")
            continue
    return output_folder


# ------------------------ 选科转换与日期处理 helpers ------------------------
def convert_selection_requirements(df):
    subject_mapping = {'物理': '物', '化学': '化', '生物': '生', '历史': '历', '地理': '地', '政治': '政',
                       '思想政治': '政'}
    df_new = df.copy()
    df_new['首选科目'] = ''
    df_new['选科要求类型'] = ''
    df_new['次选'] = ''

    # iterate rows - we selected "row" granular progress behavior
    total_rows = len(df)
    for idx, row in progress_iter(list(df.iterrows()), text="选科转换中"):
        try:
            i, r = row
            text = str(r.get('选科要求', '')).strip()
            cat = str(r.get('招生科类', '')).strip()
            subjects = [subject_mapping.get(s, s) for s in
                        re.findall(r'物理|化学|生物|历史|地理|政治|思想政治', text)]
            first = ''
            for s_full, s_short in subject_mapping.items():
                if f'首选{s_full}' in text:
                    first = s_short
            if not first:
                if '物理' in cat:
                    first = '物'
                elif '历史' in cat:
                    first = '历'
            remaining = [s for s in subjects if s != first]
            second = ''.join(remaining)
            if '不限' in text:
                req_type = '不限科目专业组'
            elif len(remaining) >= 1:
                req_type = '多门选考'
            else:
                req_type = '单科、多科均需选考'
            df_new.at[i, '首选科目'] = first
            df_new.at[i, '次选'] = second
            df_new.at[i, '选科要求类型'] = req_type
        except Exception as e:
            log(f"选科行处理失败: idx={i} -> {e}", level="warning")
            continue
    return df_new


def safe_parse_datetime(datetime_str, year):
    if pd.isna(datetime_str):
        return None
    datetime_str = str(datetime_str).strip()
    if not re.search(r'(^|\D)\d{4}(\D|$)', datetime_str):
        datetime_str = f"{year}年{datetime_str}"
    patterns = [(r'(\d{4})年(\d{1,2})月(\d{1,2})日(\d{1,2}):(\d{1,2})', '%Y年%m月%d日%H:%M'),
                (r'(\d{4})年(\d{1,2})月(\d{1,2})日', '%Y年%m月%d日'),
                (r'(\d{4})-(\d{1,2})-(\d{1,2})', '%Y-%m-%d'),
                (r'(\d{4})/(\d{1,2})/(\d{1,2})', '%Y/%m/%d')]
    for pattern, fmt in patterns:
        try:
            dt = datetime.strptime(datetime_str, fmt)
            return dt
        except Exception:
            continue
    return None


def process_date_range(date_str, year):
    try:
        if not isinstance(date_str, str) or not date_str.strip():
            return date_str, "", ""

        raw = date_str
        text = date_str.strip()

        # =========================
        # 1️⃣ 统一所有中文表达
        # =========================
        text = text.replace("—", "-")
        text = text.replace("–", "-")
        text = text.replace("至", "-")
        text = text.replace("开始", "")
        text = text.replace("止", "")
        text = text.replace("，", "")
        text = text.replace(",", "")
        text = text.replace("：", ":")
        text = text.replace("时", ":00")
        text = re.sub(r"\s+", "", text)

        # =========================
        # 2️⃣ 同月区间：6月12-18日
        # =========================
        m_same = re.fullmatch(r"(\d{1,2})月(\d{1,2})-(\d{1,2})日", text)
        if m_same:
            m, d1, d2 = m_same.groups()
            start_dt = datetime(year, int(m), int(d1), 0, 0)
            end_dt = datetime(year, int(m), int(d2), 23, 59)
            return raw, start_dt.strftime("%Y:%m:%d %H:%M:%S"), end_dt.strftime("%Y:%m:%d %H:%M:%S")

        # =========================
        # 3️⃣ 单日：7月12日
        # =========================
        m_single = re.fullmatch(r"(\d{1,2})月(\d{1,2})日", text)
        if m_single:
            m, d = m_single.groups()
            start_dt = datetime(year, int(m), int(d), 0, 0)
            end_dt = datetime(year, int(m), int(d), 23, 59)
            return raw, start_dt.strftime("%Y:%m:%d %H:%M:%S"), end_dt.strftime("%Y:%m:%d %H:%M:%S")

        # =========================
        # 4️⃣ 解析所有日期
        # =========================
        date_matches = re.findall(r"(\d{1,2})月(\d{1,2})日", text)

        # =========================
        # 5️⃣ 解析所有时间
        # =========================
        time_matches = re.findall(r"\d{1,2}:\d{2}", text)

        # =========================
        # 6️⃣ 6月7日9:00-11:30（同日时间段）
        # =========================
        same_day_time = re.fullmatch(r"(\d{1,2})月(\d{1,2})日(\d{1,2}:\d{2})-(\d{1,2}:\d{2})", text)
        if same_day_time:
            m, d, t1, t2 = same_day_time.groups()
            start_dt = datetime.strptime(f"{year}-{m}-{d} {t1}", "%Y-%m-%d %H:%M")
            end_dt = datetime.strptime(f"{year}-{m}-{d} {t2}", "%Y-%m-%d %H:%M")
            if end_dt < start_dt:
                end_dt += timedelta(days=1)
            return raw, start_dt.strftime("%Y:%m:%d %H:%M:%S"), end_dt.strftime("%Y:%m:%d %H:%M:%S")

        # =========================
        # 7️⃣ 两个完整日期 + 时间
        # =========================
        if len(date_matches) == 2 and len(time_matches) >= 2:
            (m1, d1), (m2, d2) = date_matches[:2]
            t1, t2 = time_matches[:2]

            start_dt = datetime.strptime(f"{year}-{m1}-{d1} {t1}", "%Y-%m-%d %H:%M")
            end_dt = datetime.strptime(f"{year}-{m2}-{d2} {t2}", "%Y-%m-%d %H:%M")

            return raw, start_dt.strftime("%Y:%m:%d %H:%M:%S"), end_dt.strftime("%Y:%m:%d %H:%M:%S")

        # =========================
        # 8️⃣ 两个完整日期无时间
        # =========================
        if len(date_matches) == 2 and len(time_matches) == 0:
            (m1, d1), (m2, d2) = date_matches[:2]
            start_dt = datetime(year, int(m1), int(d1), 0, 0)
            end_dt = datetime(year, int(m2), int(d2), 23, 59)
            return raw, start_dt.strftime("%Y:%m:%d %H:%M:%S"), end_dt.strftime("%Y:%m:%d %H:%M:%S")

        return raw, "格式错误", "格式错误"

    except Exception:
        return date_str, "格式错误", "格式错误"
# ------------------------ tab1函数 ------------------------
# ================= 工具函数（最终稳定版：只保留最佳表） =================

import pandas as pd
import requests
from bs4 import BeautifulSoup
from io import StringIO, BytesIO
from datetime import datetime
import streamlit as st
import re


# ================= 日志 =================
def log(msg, level="info"):
    s = f"{datetime.now().strftime('%H:%M:%S')} [{level.upper()}] {msg}"
    print(s)
    if "recent_logs" not in st.session_state:
        st.session_state.recent_logs = []
    st.session_state.recent_logs.append(s)


# ================= 列名唯一 =================
def make_columns_unique(df):
    cols = list(df.columns)
    new_cols = []
    counter = {}

    for col in cols:
        col = str(col).strip()

        if col == "" or col.lower().startswith("unnamed"):
            col = "空列"

        if col not in counter:
            counter[col] = 0
            new_cols.append(col)
        else:
            counter[col] += 1
            new_cols.append(f"{col}_{counter[col]}")

    df.columns = new_cols
    return df


# ================= 清理列 =================
def clean_columns(df):
    df = df.copy()

    df = df.dropna(axis=1, how="all")

    drop_cols = []
    for col in df.columns:
        if "Unnamed" in str(col):
            if df[col].isna().mean() > 0.9:
                drop_cols.append(col)

    df = df.drop(columns=drop_cols, errors="ignore")

    return df


# ================= 表头压平 =================
def flatten_columns(df):
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            "_".join([str(c) for c in col if str(c) != "nan"]).strip()
            for col in df.columns.values
        ]
    else:
        df.columns = [str(c).strip() for c in df.columns]

    return df


# ================= pandas解析 =================
def parse_tables_pandas(html):
    tables = []

    for header in [[0], [0,1], [0,1,2]]:
        try:
            t = pd.read_html(StringIO(html), header=header)
            for df in t:
                df = flatten_columns(df)
                df = clean_columns(df)
                df = make_columns_unique(df)
                tables.append(df)
        except:
            continue

    return tables


# ================= DOM解析 =================
def parse_tables_dom(html):
    soup = BeautifulSoup(html, "html.parser")
    tables = []

    for table in soup.find_all("table"):
        rows = []
        for tr in table.find_all("tr"):
            cols = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if cols:
                rows.append(cols)

        if len(rows) >= 2:
            df = pd.DataFrame(rows)
            df = make_columns_unique(df)
            tables.append(df)

    return tables


# ================= 表去重 =================
def table_hash(df):
    try:
        df2 = df.copy().fillna("")
        df2 = df2.sort_index(axis=1)
        return hash(pd.util.hash_pandas_object(df2, index=True).sum())
    except:
        return hash(str(df.values))


def deduplicate_tables(tables):
    seen = set()
    result = []

    for df in tables:
        h = table_hash(df)
        if h not in seen:
            seen.add(h)
            result.append(df)

    return result


# ================= ⭐ 表质量评分（核心） =================
def score_table(df):
    cols = list(df.columns)

    score = 0

    # ❌ 垃圾列名
    score -= sum(str(c).startswith("Unnamed") for c in cols) * 5
    score -= sum(str(c).isdigit() for c in cols) * 3

    # ❌ 列太少
    score -= max(0, 3 - len(cols)) * 2

    # ✅ 招生关键词
    score += sum(any(k in str(c) for k in ["专业","分","位","批次"]) for c in cols) * 3

    # ✅ 数据量
    score += df.notna().sum().sum() * 0.001

    return score


# ================= ⭐ 只保留最佳表 =================
def pick_best_table(tables):
    if not tables:
        return []

    scored = [(score_table(df), df) for df in tables]
    scored.sort(reverse=True, key=lambda x: x[0])

    best_score, best_df = scored[0]

    log(f"表评分: {[round(s,2) for s,_ in scored]}")

    # 阈值保护（防止全是垃圾）
    if best_score < 0:
        return []

    return [best_df]


# ================= 列排序 =================
def reorder_columns(df):
    priority = [
        "学校名称", "专业名称", "省份",
        "批次", "科类", "最低分", "位次"
    ]

    cols = list(df.columns)

    ordered = [c for c in priority if c in cols]
    rest = [c for c in cols if c not in ordered]

    return df[ordered + rest]


# ================= 填充 =================
def smart_fill(df, manual_cols=None):
    df = df.copy()

    if not manual_cols:
        return df, []

    valid_cols = [c for c in manual_cols if c in df.columns]

    log(f"填充列: {valid_cols}")

    for col in valid_cols:
        df[col] = df[col].replace("", None)
        df[col] = df[col].ffill()

    return df, valid_cols


# ================= 抓取 =================
def smart_fetch(url):
    log(f"抓取: {url}")
    tables = []

    try:
        r = requests.get(url, timeout=15)
        r.encoding = r.apparent_encoding
        html = r.text
    except Exception as e:
        log(f"请求失败: {e}", "error")
        return []

    tables += parse_tables_pandas(html)
    tables += parse_tables_dom(html)

    log(f"原始表数: {len(tables)}")

    tables = deduplicate_tables(tables)
    log(f"去重后: {len(tables)}")

    tables = pick_best_table(tables)
    log(f"保留表数: {len(tables)}")

    return tables


# ================= sheet名 =================
def clean_sheet_name(url, idx):
    name = re.sub(r"https?://", "", url)
    name = name.split("/")[0]
    name = re.sub(r"[\\/*?:\[\]]", "", name)
    return f"{idx}_{name}"[:31]


# ================= 主函数 =================
def scrape_table(url_list, fill_cols=None, progress_callback=None):
    output = BytesIO()

    total = len(url_list)

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

        for i, url in enumerate(url_list):

            try:
                tables = smart_fetch(url)

                if not tables:
                    log(f"{url} 未抓到有效表", "warning")
                    continue

                df = pd.concat(tables, ignore_index=True, sort=False)

                df = clean_columns(df)
                df = make_columns_unique(df)

                df, used_cols = smart_fill(df, fill_cols)

                df = reorder_columns(df)

                sheet_name = clean_sheet_name(url, i+1)

                df.to_excel(writer, index=False, sheet_name=sheet_name)

                log(f"写入 {sheet_name} 行数: {len(df)}")

            except Exception as e:
                log(f"{url} 失败: {e}", "error")

            if progress_callback:
                progress_callback(i+1, total)

    output.seek(0)
    return output
# ------------------------ Streamlit UI ------------------------
st.title("🧰 综合处理工具箱 - 完整版（带进度条 & 日志）")
tab1, tab2, tab3, tab4, tab5, tab6 , tab7= st.tabs([
    "网页表格抓取",
    "网页图片下载",
    "Excel日期处理",
    "分数匹配",
    "学业桥-高考专业分数据转换",
    "院校分提取",
    "学业桥-招生计划模板转换"

])

# side: logs
with st.sidebar.expander("运行日志（最新）", expanded=True):
    for line in st.session_state.recent_logs[-200:]:
        st.text(line)

# ------------------------ Tab1: 网页表格抓取 ------------------------
with tab1:
    st.subheader("网页表格抓取（手动填充版）🔥")
    st.markdown("支持：复杂网页抓取 + 多表合并 + 手动填充")

    # 初始化日志
    if "recent_logs" not in st.session_state:
        st.session_state.recent_logs = []

    urls_text = st.text_area(
        "输入网页URL列表（每行一个）",
        height=160,
        placeholder="https://xxx\nhttps://xxx"
    )

    # ===== 手动填充（选填）=====
    manual_cols_input = st.text_input(
        "手动填充列（选填）",
        placeholder="例如: 省份,批次,科类（支持中文逗号）"
    )

    # ===== 高级选项 =====
    with st.expander("🔧 高级选项", expanded=False):
        debug_mode = st.checkbox("显示日志", value=True)

    if st.button("🚀 开始抓取", type="primary"):

        # ===== URL处理 =====
        url_list = [u.strip() for u in urls_text.splitlines() if u.strip()]

        if not url_list:
            st.warning("⚠️ 请先输入URL")
            st.stop()

        # ===== 填充列处理（唯一正确写法）=====
        import re

        if manual_cols_input:
            fill_cols = re.split(r"[,\，]", manual_cols_input)
            fill_cols = [c.strip() for c in fill_cols if c.strip()]
            if not fill_cols:
                fill_cols = None
        else:
            fill_cols = None

        # ===== 提示当前策略 =====
        if fill_cols:
            st.info(f"🧠 当前填充列: {fill_cols}")
        else:
            st.info("🧠 未设置填充列（将不进行填充）")

        # ===== UI组件 =====
        status_placeholder = st.empty()
        progress_bar = st.progress(0)

        status_placeholder.info(f"🔄 正在抓取 {len(url_list)} 个网页...")

        def progress_callback(i, total):
            progress_bar.progress(i / total)

        # ===== 执行 =====
        with st.spinner("抓取中，请稍候..."):
            output = scrape_table(
                url_list,
                fill_cols=fill_cols,
                progress_callback=progress_callback
            )

        # ===== 结果 =====
        if output:
            status_placeholder.success("✅ 抓取完成！")

            total_size = len(output.getvalue()) / 1024

            st.success(
                f"""
                **抓取结果**
                - 文件大小: {total_size:.1f} KB
                - 已合并所有表格 ✅
                - 已处理复杂结构 ✅
                """
            )

            st.download_button(
                "📥 下载Excel",
                data=output.getvalue(),
                file_name=f"网页抓取_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                type="primary"
            )

        else:
            status_placeholder.error("❌ 未抓取到表格数据")

    # ===== 日志 =====
    if debug_mode:
        with st.expander("📋 处理日志", expanded=False):
            logs = st.session_state.recent_logs[-50:]
            for log_entry in logs:
                st.text(log_entry)
# ------------------------ Tab 2: 网页图片下载 ------------------------
with tab2:
    st.subheader("网页图片下载")
    st.markdown("同时抓取多个链接的图片，适用于招生快讯类图片类的录取数据")
    urls_text2 = st.text_area("输入网页URL列表（每行一个）", height=160, key="img_urls")

    # 显示当前工作目录信息
    st.info(f"当前工作目录: `{os.getcwd()}`")
    st.info(f"临时文件目录: `{tempfile.gettempdir()}`")

    col1, col2 = st.columns([1, 3])
    with col1:
        if st.button("下载图片", key="img_download"):
            url_list = [u.strip() for u in urls_text2.splitlines() if u.strip()]
            if not url_list:
                st.warning("请先输入有效URL列表")
            else:
                try:
                    output_dir, files, errors = download_images_from_urls(url_list)

                    # 显示下载结果
                    st.success(f"✅ 完成！共下载 {len(files)} 张图片")
                    st.success(f"📁 保存到: `{output_dir}`")

                    # 显示文件列表
                    if files:
                        st.subheader("📄 下载的文件列表:")

                        # 创建ZIP下载
                        zip_buffer = create_zip_download(files)
                        st.download_button(
                            label="📦 下载所有图片(ZIP)",
                            data=zip_buffer.getvalue(),
                            file_name="downloaded_images.zip",
                            mime="application/zip"
                        )

                        # 显示文件详情和预览
                        for i, file_path in enumerate(files, 1):
                            file_name = os.path.basename(file_path)
                            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0

                            col1, col2 = st.columns([3, 1])
                            with col1:
                                st.write(f"{i}. **{file_name}** ({file_size} bytes)")
                            with col2:
                                # 单个文件下载
                                with open(file_path, 'rb') as f:
                                    st.download_button(
                                        f"下载{i}",
                                        f.read(),
                                        file_name=file_name,
                                        key=f"single_{i}"
                                    )

                            # 图片预览
                            try:
                                st.image(file_path, caption=file_name, width=300)
                            except Exception as e:
                                st.write(f"预览失败: {e}")

                    if errors:
                        st.warning(f"有 {len(errors)} 个错误:")
                        for error in errors[-5:]:
                            st.error(error)

                except Exception as e:
                    log(f"下载图片失败: {e}\n{traceback.format_exc()}", level="error")
                    st.error(f"下载图片出错: {e}")

# ------------------------ Tab 3: Excel日期处理 ------------------------
with tab3:
    st.subheader("Excel日期处理")
    uploaded_file2 = st.file_uploader("上传Excel文件", type=["xlsx", "xls"], key="date_excel")
    year = st.number_input("年份（用于补全）", value=datetime.now().year, key="date_year")
    date_col = st.text_input("日期列名", value="日期", key="date_col")

    if uploaded_file2:
        try:
            df2 = pd.read_excel(uploaded_file2)
            st.write("原始数据预览", df2.head())
            if st.button("处理日期", key="date_btn"):
                try:
                    start_times = []
                    end_times = []
                    originals = []
                    # row-by-row processing (you selected 'row' granular mode)
                    for d in progress_iter(list(df2[date_col]), text="日期处理中"):
                        orig, start, end = process_date_range(d, int(year))
                        originals.append(orig)
                        start_times.append(start)
                        end_times.append(end)
                    df2_result = df2.copy()
                    insert_at = df2_result.columns.get_loc(date_col) + 1
                    df2_result.insert(insert_at, '开始时间', start_times)
                    df2_result.insert(insert_at + 1, '结束时间', end_times)
                    st.write("处理结果预览", df2_result.head())
                    towrite2 = BytesIO()
                    df2_result.to_excel(towrite2, index=False)
                    towrite2.seek(0)
                    st.download_button("下载日期处理结果Excel", data=towrite2.getvalue(), file_name="日期处理结果.xlsx")
                    st.success("日期处理完成")
                except Exception as e:
                    log(f"日期处理失败: {e}\n{traceback.format_exc()}", level="error")
                    st.error("日期处理出错，详情见日志")
        except Exception as e:
            log(f"读取上传文件失败: {e}", level="error")
            st.error("无法读取上传的 Excel 文件")

# =====================================================
# ======================= TAB 4 =======================
# =====================================================
with tab4:
    st.header("🎓 招生计划 & 分数表 智能匹配工具")

    st.markdown("""
    ### 📋 功能说明
    将招生计划和分数表进行智能匹配，匹配成功后自动转换为专业分模板格式。
    
    加入了事先校验上传的计划表和分数表是否有重复数据的逻辑（同一学校科类层次批次招生类型专业），避免有未处理到的重复数据导致分数匹配错误
    
    
    加入了根据 学校+省份+科类+层次+专业+招生类型 匹配批次的逻辑，分数表可以不填批次，如果批次唯一则自动填充，如果批次不唯一则提示
    
    **匹配规则**（计划表和分数表需保持一致）:
    - 学校
    - 省份  
    - 科类
    - 批次（可以根据计划表匹配，可不填）
    - 层次
    - 专业
    - 招生类型
    **需要上传以下2个文件：**
    1. **分数表** - 整理好的需要入库的分数（可按下载的模板整理）
    2. **计划表** - 从高考数据库导出的计划数据

    3.**重复匹配人工确认区-重复数据是指一条计划有两条对应分数或两条计划对应一条分数**

    **⚠️ 重要提示：**
    处理完的数据需要抽查一下，确保匹配准确！""")

    # ================= 匹配配置 =================
    MATCH_KEYS = ["学校", "省份", "科类", "层次", "批次", "招生类型", "专业"]

    # ⚠ 只用于【人工重复匹配区-分数】候选项展示
    DISPLAY_FIELDS = [
        "学校",
        "省份",
        "科类",
        "批次",
        "专业",
        "备注",
        "招生类型"
    ]

    TEXT_COLUMNS = {"专业组代码", "招生代码", "专业代码"}

    # ================= 分数表批次自动补全 =================
    BATCH_FILL_KEYS = [
        "学校",
        "省份",
        "科类",
        "层次",
        "专业",
        "招生类型"
    ]


    def build_fill_batch_key(df: pd.DataFrame) -> pd.Series:
        cols = []

        for c in BATCH_FILL_KEYS:
            norm_col = c + "_norm"

            if norm_col in df.columns:
                cols.append(norm_col)

        if not cols:
            return pd.Series([""] * len(df), index=df.index)

        return (
            df[cols]
            .fillna("")
            .astype(str)
            .agg("||".join, axis=1)
        )


    # ================= 工具函数 =================
    def normalize(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        for col in MATCH_KEYS:
            if col in df.columns:
                df[col + "_norm"] = (
                    df[col]
                    .fillna("")  # ⭐ 填充空值
                    .astype(str)  # ⭐ 强制转为字符串
                    .str.strip()
                    .str.replace("\u3000", "")
                    .str.lower()
                )

        if "层次_norm" in df.columns:
            df["层次_norm"] = df["层次_norm"].replace({
                "专科": "专科(高职)"
            })

        return df



            

    #------------------------校验计划表 分数表是否有重复数据 学校 省份 科类  批次  招生类型  专业 层次---------------#


    def check_duplicates(df: pd.DataFrame, name: str) -> pd.DataFrame:
        # ⭐ 已加入“层次”
        check_keys = ["学校", "省份", "科类", "层次", "批次", "招生类型", "专业"]

        exist_keys = [k for k in check_keys if k in df.columns]
        if len(exist_keys) < len(check_keys):
            st.warning(f"⚠ {name} 缺少部分校验字段，已跳过重复校验")
            return pd.DataFrame()

        # 找重复
        dup_mask = df.duplicated(subset=exist_keys, keep=False)
        dup_df = df[dup_mask].copy()

        if dup_df.empty:
            st.success(f"✅ {name} 未发现重复数据")
            return dup_df

        # ⭐ 关键修复：避免 NaN 导致分组失效
        dup_df[exist_keys] = dup_df[exist_keys].fillna("")

        dup_df["_重复组"] = dup_df.groupby(exist_keys, dropna=False).ngroup()
        group_count = dup_df["_重复组"].nunique()

        st.error(f"❌ {name} 存在重复数据：共 {len(dup_df)} 条，{group_count} 组")

        for gid, group in dup_df.groupby("_重复组"):
            first = group.iloc[0]

            title = " | ".join([
                f"{k}:{first.get(k, '')}" for k in exist_keys
            ])

            with st.expander(f"🔁 重复组 {gid + 1}（{len(group)} 条）｜{title}"):
                st.dataframe(
                    group.drop(columns=["_重复组"]),
                    use_container_width=True,
                    height=300  # ⭐ 滚动关键
                )
        return dup_df


    def build_key(df: pd.DataFrame) -> pd.Series:
        norm_cols = [c + "_norm" for c in MATCH_KEYS if c + "_norm" in df.columns]
        if not norm_cols:  # ⚠ 兜底：如果没有_norm列
            return pd.Series([""] * len(df), index=df.index)
        return df[norm_cols].fillna("").astype(str).agg("||".join, axis=1)


    def safe_text(v):
        return str(v) if pd.notna(v) and str(v).strip() else ""


    def clean_code_text(v):
        if pd.isna(v):
            return ""
        s = str(v).strip()
        if s.startswith("^"):
            s = s[1:]
        return s


    # ================= 选科解析 =================
    def calc_first_subject(kl: str) -> str:
        if not isinstance(kl, str):
            return ""
        if "历史" in kl:
            return "历"
        if "物理" in kl:
            return "物"
        return ""


    def parse_subject_requirement(require_text: str, kl: str) -> tuple[str, str]:
        # ===== 兜底方案（核心修改点 1）=====
        # 原字段为空 / NaN / 空白 / nan字符串 → 绝对不生成选科结果
        if require_text is None:
            return "", ""

        s = str(require_text).strip()

        if s == "" or s.lower() in {"nan", "none"}:
            return "", ""

        # ===== 原有逻辑，完全保留 =====
        if "文科" in kl or "理科" in kl:
            return "", ""

        text = (
            s.replace(" ", "")
            .replace("　", "")
            .replace("，", "")
            .replace(",", "")
            .replace("、", "")

        )
        subject_map = [
            ("思想政治", "政"),
            ("政治", "政"),

            ("物理", "物"),
            ("历史", "历"),
            ("化学", "化"),
            ("生物", "生"),
            ("地理", "地"),
        ]


        def extract_all(s: str) -> str:
            res = []
            for k, v in subject_map:
                if k in s and v not in res:
                    res.append(v)
            return "".join(res)

        def extract_after_reselect(s: str) -> str:
            if "再选" in s:
                s = s.split("再选", 1)[1]
            return extract_all(s)

        if "不限" in text:
            return "不限科目专业组", ""

        must_keywords = ["必选", "均需", "全部", "全选", "均须", "3科必选"]
        multi_keywords = ["或", "/", "任选", "选一", "至少", "其中", "之一","2选1","3选1"]

        is_must = any(k in text for k in must_keywords)
        is_multi = any(k in text for k in multi_keywords)

        req_type = "多门选考" if is_multi else "单科、多科均需选考"

        # 一律先抽取全部科目
        second = extract_all(text)

        # ⭐ 核心规则：只要是物理 / 历史类，必须剔除首选
        if "物理" in kl:
            second = second.replace("物", "")
        elif "历史" in kl:
            second = second.replace("历", "")

        return req_type, second


    # ================= 核心合并 =================
    def merge_plan_score(plan_row: pd.Series, score_row: pd.Series) -> dict:
        # ===== 兜底方案（核心修改点 2，双保险）=====
        raw_req = plan_row.get("专业选科要求(新高考专业省份)", "")

        if not isinstance(raw_req, str) or not raw_req.strip():
            select_req, second_req = "", ""
        else:
            select_req, second_req = parse_subject_requirement(
                raw_req,
                plan_row.get("科类", "")
            )

        enroll_count = score_row.get("招生人数", "")
        if pd.isna(enroll_count) or enroll_count == "":
            enroll_count = plan_row.get("招生人数", "")

        level = plan_row.get("层次", "")
        if level == "专科":
            level = "专科(高职)"

        return {
            "学校名称": plan_row.get("学校", ""),
            "省份": plan_row.get("省份", ""),
            "招生专业": plan_row.get("专业", ""),
            "专业方向（选填）": plan_row.get("专业方向", ""),
            "专业备注（选填）": plan_row.get("备注", ""),
            "层次": level,
            "招生科类": plan_row.get("科类", ""),
            "招生批次": plan_row.get("批次", ""),
            "招生类型（选填）": plan_row.get("招生类型", ""),
            "最高分": score_row.get("最高分", ""),
            "最低分": score_row.get("最低分", ""),
            "平均分": score_row.get("平均分", ""),
            "最低分位次": score_row.get("最低分位次", ""),
            "招生人数": enroll_count,
            "数据来源": "学校官网",
            "专业组代码": clean_code_text(plan_row.get("专业组代码", "")),
            "首选科目": calc_first_subject(plan_row.get("科类", "")),
            "选科要求": select_req,
            "次选": second_req,
            "专业代码": clean_code_text(plan_row.get("专业代码", "")),
            "招生代码": clean_code_text(plan_row.get("招生代码", "")),
            "最低分数区间低": "",
            "最低分数区间高": "",
            "最低分数区间位次低": "",
            "最低分数区间位次高": "",
            "录取人数": score_row.get("录取人数", ""),
        }


    def diff_fields(df: pd.DataFrame, fields: list[str]) -> set[str]:
        diffs = set()
        for f in fields:
            if f in df.columns and df[f].nunique(dropna=False) > 1:
                diffs.add(f)
        return diffs


    def clear_cache():
        st.session_state.chosen = {}
        st.session_state.expanded = {}


    # ================= 分数表模板下载 =================
    st.subheader("📥 分数表导入模板")

    template_cols = [
        "学校", "省份", "科类", "层次", "批次", "专业", "备注", "招生类型",
        "最高分", "最低分", "平均分", "最低分位次", "招生人数", "录取人数"
    ]

    template_df = pd.DataFrame(columns=template_cols)
    buf = BytesIO()
    template_df.to_excel(buf, index=False)
    buf.seek(0)

    st.download_button(
        "⬇ 下载【分数表】Excel模板",
        data=buf,
        file_name="分数表导入模板.xlsx"
    )

    # ================= 数据上传 =================
    st.info("此功能需要上传计划表和分数表。请使用上方上传控件。")
    plan_file = st.file_uploader("📘 上传【计划表】Excel", type=["xls", "xlsx"], key="plan_file_4")
    score_file = st.file_uploader("📙 上传【分数表】Excel", type=["xls", "xlsx"], key="score_file_4")

    # 初始化session state
    if "chosen" not in st.session_state:
        st.session_state.chosen = {}
    if "expanded" not in st.session_state:
        st.session_state.expanded = {}

    if plan_file and score_file:
        try:
            # 读取原始数据
            raw_plan_df = pd.read_excel(plan_file)
            raw_score_df = pd.read_excel(score_file)

            # 数据标准化
            plan_df = normalize(raw_plan_df)
            score_df = normalize(raw_score_df)

            # =====================================================
            # 批次字段统一清洗（非常重要）
            # =====================================================

            plan_df["批次"] = (
                plan_df["批次"]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            score_df["批次"] = (
                score_df["批次"]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            # =====================================================
            # 分数表批次自动补全
            # =====================================================

            # 构建key（不含批次）
            plan_df["_fill_batch_key"] = build_fill_batch_key(plan_df)
            score_df["_fill_batch_key"] = build_fill_batch_key(score_df)

            # 同key对应的批次
            plan_batch_map = (
                plan_df.groupby("_fill_batch_key")["批次"]
                .apply(
                    lambda x: sorted(
                        set(
                            str(i).strip()
                            for i in x
                            if pd.notna(i) and str(i).strip()
                        )
                    )
                )
                .to_dict()
            )

            # 冲突记录
            batch_conflict_rows = []

            # 遍历分数表
            for idx, row in score_df.iterrows():

                # 已有批次直接跳过
                current_batch = str(row.get("批次", "")).strip()

                if current_batch:
                    continue

                key = row["_fill_batch_key"]

                # 计划表里的批次
                plan_batches = plan_batch_map.get(key, [])

                # 没找到计划
                if len(plan_batches) == 0:
                    continue

                # ================= 只有一个批次 -> 自动补 =================
                if len(plan_batches) == 1:

                    score_df.at[idx, "批次"] = plan_batches[0]

                # ================= 多个批次 -> 记录冲突 =================
                else:

                    batch_conflict_rows.append({
                        "学校": row.get("学校", ""),
                        "省份": row.get("省份", ""),
                        "科类": row.get("科类", ""),
                        "层次": row.get("层次", ""),
                        "专业": row.get("专业", ""),
                        "招生类型": row.get("招生类型", ""),
                        "计划表存在批次": "、".join(plan_batches)
                    })

            # ⭐ 非常重要：补完批次后重新normalize
            score_df = normalize(score_df)

            # =====================================================
            # 强制保留补完后的批次
            # =====================================================

            score_df["批次"] = (
                score_df["批次"]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            # =====================================================
            # 下载补全后的分数表
            # =====================================================

            if batch_conflict_rows:

                st.warning(
                    f"⚠ 存在 {len(batch_conflict_rows)} 条无法自动补全批次的数据"
                )

                conflict_df = pd.DataFrame(batch_conflict_rows)

                st.dataframe(
                    conflict_df,
                    use_container_width=True,
                    height=400
                )

            else:

                st.success("✅ 分数表批次自动补全完成")

            # =====================================================
            # 导出当前已补全的分数表
            # =====================================================

            output_score = BytesIO()

            export_score_df = score_df.copy()

            # 删除内部字段
            drop_cols = [
                c for c in export_score_df.columns
                if c.endswith("_norm")
                   or c.startswith("_")
            ]

            export_score_df = export_score_df.drop(
                columns=drop_cols,
                errors="ignore"
            )

            export_score_df.to_excel(
                output_score,
                index=False
            )

            output_score.seek(0)


            st.write(
                score_df[
                    ["学校", "专业", "招生类型", "批次"]
                ].head(30)
            )

            st.download_button(
                "⬇ 下载【批次已自动补全】分数表",
                data=output_score,
                file_name="分数表_批次已补全.xlsx"
            )

            st.info(
                "✅ 已自动使用补全后的批次继续进行正式匹配"
            )

    

            # ================= 重复校验 =================
            st.subheader("🔍 源数据重复校验")

            plan_dup = check_duplicates(plan_df, "计划表")
            score_dup = check_duplicates(score_df, "分数表")



            st.success(f"✅ 成功读取数据！计划表: {len(plan_df)} 行，分数表: {len(score_df)} 行")

            # ================= 选科要求字段清洗（仅此一列） =================
            SUBJECT_COL = "专业选科要求(新高考专业省份)"

            if SUBJECT_COL in plan_df.columns:
                plan_df[SUBJECT_COL] = (
                    plan_df[SUBJECT_COL]
                    .astype(str)
                    .str.strip()
                    .str.replace(r"^\^", "", regex=True)  # ⭐ 关键：去掉开头 ^
                    .replace({"nan": "", "None": ""})
                )

            if "专业选科要求(新高考专业省份)" not in plan_df.columns:
                plan_df["专业选科要求(新高考专业省份)"] = ""

            # 检查必要字段
            missing_in_plan = [k for k in MATCH_KEYS if k not in plan_df.columns]
            missing_in_score = [k for k in MATCH_KEYS if k not in score_df.columns]

            if missing_in_plan:
                st.error(f"❌ 计划表缺少字段：{missing_in_plan}")
                st.stop()
            if missing_in_score:
                st.error(f"❌ 分数表缺少字段：{missing_in_score}")
                st.stop()

            # 构建匹配键
            plan_df["_key"] = build_key(plan_df)
            score_df["_key"] = build_key(score_df)
            plan_key_count = plan_df["_key"].value_counts()
            score_groups = score_df.groupby("_key")



            # ================= 匹配（按key分组版） =================
            unique_rows = []
            duplicate_rows = []
            unmatched_rows = []

            plan_groups = plan_df.groupby("_key")

            for key, plan_group in plan_groups:
                if key not in score_groups.groups:
                    # 所有计划都未匹配
                    for _, plan_row in plan_group.iterrows():
                        unmatched_rows.append(plan_row)
                else:
                    score_group = score_groups.get_group(key)

                    plan_count = len(plan_group)
                    score_count = len(score_group)

                    if plan_count == 1 and score_count == 1:
                        # 真正1对1
                        row = merge_plan_score(plan_group.iloc[0], score_group.iloc[0])
                        unique_rows.append(row)


                    else:
                        # ⭐ 按“计划”进入人工区（核心）
                        duplicate_rows.append({
                            "plans": plan_group,
                            "scores": score_group
                        })

            # ================= 统计 =================
            st.success(
                f"✅ 唯一匹配：{len(unique_rows)} 条 ｜ "
                f"⚠ 重复匹配：{len(duplicate_rows)} 条 ｜ "
                f"❌ 未匹配：{len(unmatched_rows)} 条"
            )

            # ================= 重复匹配 =================
            st.header("⚠ 重复匹配人工确认区")

            # 需要确认的总计划数
            total_dup = sum(len(item["plans"]) for item in duplicate_rows)

            # 已确认数量
            confirmed = len(st.session_state.chosen)

            # 进度条
            progress = 1.0 if total_dup == 0 else confirmed / total_dup

            st.progress(progress)

            st.caption(
                f"已确认 {confirmed} / {total_dup} 条计划（{int(progress * 100)}%）"
            )

            # 遍历重复组
            for group_idx, item in enumerate(duplicate_rows):

                plan_group = item["plans"]

                candidates = item["scores"]

                first = plan_group.iloc[0]

                title_fields = [
                    "学校",
                    "省份",
                    "科类",
                    "批次",
                    "专业",
                    "备注",
                    "招生类型"
                ]

                title_parts = []

                for field in title_fields:

                    value = first.get(field, "")

                    if pd.notna(value) and str(value).strip():
                        title_parts.append(str(value))

                # ================= 重复类型 =================
                if len(plan_group) > 1 and len(candidates) == 1:

                    tag = "⚠ 多计划→1分数"

                elif len(plan_group) == 1 and len(candidates) > 1:

                    tag = "⚠ 1计划→多分数"

                else:

                    tag = "⚠ 多对多"

                # ================= 标题字段 =================
                title_fields = [
                    "学校",
                    "省份",
                    "科类",
                    "批次",
                    "专业",
                    "备注",
                    "招生类型"
                ]

                title_parts = []

                for field in title_fields:

                    value = first.get(field, "")

                    if pd.notna(value) and str(value).strip():
                        title_parts.append(str(value))

                title = (
                        f"{group_idx + 1}. "
                        f"{tag} | "
                        + " | ".join(title_parts)
                )

                with st.expander(title, expanded=False):

                    # 当前已经被使用的分数
                    used_scores = {
                        v for v in st.session_state.chosen.values()
                        if v != "NO_SCORE"
                    }

                    # ⭐ 逐计划处理（核心）
                    for plan_idx, (_, plan_row) in enumerate(plan_group.iterrows()):

                        # 当前计划唯一key
                        choice_key = f"{group_idx}_{plan_idx}"

                        st.markdown("---")

                        # ================= 已确认 =================
                        if choice_key in st.session_state.chosen:

                            chosen_val = st.session_state.chosen[choice_key]

                            # 无分数
                            if chosen_val == "NO_SCORE":

                                st.success("✅ 已确认：无对应分数")

                            else:

                                chosen_score = score_df.loc[chosen_val]

                                st.success(
                                    f"""
                                    ✅ 已选择分数：

                                    最低分：{chosen_score.get('最低分', '')}

                                    平均分：{chosen_score.get('平均分', '')}

                                    备注：{chosen_score.get('备注', '')}

                                    招生类型：{chosen_score.get('招生类型', '')}
                                    """
                                )

                            # 重新选择
                            if st.button(
                                    "🔁 重新选择",
                                    key=f"reset_{choice_key}"
                            ):
                                del st.session_state.chosen[choice_key]

                                st.rerun()

                        # ================= 未确认 =================
                        else:

                            options = [
                                (None, "请选择对应分数"),
                                ("NO_SCORE", "🚫 无对应分数")
                            ]

                            # 候选差异字段
                            diff_cols = diff_fields(
                                candidates,
                                DISPLAY_FIELDS
                            )

                            # 构建候选项
                            for idx, r in candidates.iterrows():

                                info = []

                                for col in DISPLAY_FIELDS:

                                    if col in r and pd.notna(r[col]):

                                        # 高亮不同字段
                                        if col in diff_cols:
                                            info.append(f"🔴【{col}】{r[col]}")
                                        else:
                                            info.append(f"{col}:{r[col]}")

                                label = " | ".join(info)

                                # 已使用标记
                                if idx in used_scores:
                                    label = "🚫 已被其它计划使用 ｜ " + label

                                options.append((idx, label))

                            # 单选框
                            selected = st.radio(
                                "请选择对应分数",
                                options=options,
                                format_func=lambda x: x[1],
                                index=0,
                                key=f"radio_{choice_key}"
                            )

                            # 点击确认
                            if selected[0] is not None:

                                if st.button(
                                        "✅ 确认本计划",
                                        key=f"confirm_{choice_key}"
                                ):

                                    # 防止重复使用
                                    if (
                                            selected[0] != "NO_SCORE"
                                            and selected[0] in used_scores
                                    ):

                                        st.error("❌ 该分数已被其它计划使用")

                                    else:

                                        st.session_state.chosen[choice_key] = selected[0]

                                        st.rerun()

            if st.button("🧹 手动清理缓存（重新开始匹配）"):
                clear_cache()
                st.success("缓存已清理")
                st.rerun()

            total_need_confirm = sum(
                len(item["plans"])
                for item in duplicate_rows
            )

            all_chosen = (
                    len(st.session_state.chosen)
                    == total_need_confirm
            )

            if st.button("📥 导出最终完整数据", disabled=not all_chosen):
                final_rows = []
                final_rows.extend(unique_rows)
                used_score_indices = set()

                # ⭐ 把唯一匹配用到的分数也补进去
                for key, plan_group in plan_df.groupby("_key"):
                    if key in score_groups.groups:
                        score_group = score_groups.get_group(key)
                        if len(plan_group) == 1 and len(score_group) == 1:
                            used_score_indices.add(score_group.index[0])

                for group_idx, item in enumerate(duplicate_rows):

                    plan_group = item["plans"]

                    # ⭐ 逐计划导出
                    for plan_idx, (_, plan_row) in enumerate(plan_group.iterrows()):

                        choice_key = f"{group_idx}_{plan_idx}"

                        score_idx = st.session_state.chosen[choice_key]

                        # 无分数
                        if score_idx == "NO_SCORE":

                            score_row = pd.Series(dtype=object)

                        else:

                            used_score_indices.add(score_idx)

                            score_row = score_df.loc[score_idx]

                        # 生成最终结果
                        final_rows.append(
                            merge_plan_score(plan_row, score_row)
                        )

                final_df = pd.DataFrame(final_rows)
                unmatched_formatted = []

                for plan_row in unmatched_rows:
                    empty_score = pd.Series(dtype=object)  # ⭐ 空分数行
                    unmatched_formatted.append(
                        merge_plan_score(plan_row, empty_score)
                    )

                unmatched_df = pd.DataFrame(unmatched_formatted)

                # ===== ⭐ 未匹配分数（就放这里）=====
                unused_score_rows = []

                for _, score_row in score_df[~score_df.index.isin(used_score_indices)].iterrows():
                    fake_plan = pd.Series({
                        "学校": score_row.get("学校", ""),
                        "省份": score_row.get("省份", ""),
                        "专业": score_row.get("专业", ""),
                        "专业方向": score_row.get("专业方向", ""),
                        "备注": score_row.get("备注", ""),
                        "层次": score_row.get("层次", ""),
                        "科类": score_row.get("科类", ""),
                        "批次": score_row.get("批次", ""),
                        "招生类型": score_row.get("招生类型", ""),
                        "招生人数": score_row.get("招生人数", ""),
                    })

                    unused_score_rows.append(
                        merge_plan_score(fake_plan, score_row)
                    )

                unused_score_df = pd.DataFrame(unused_score_rows)


                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    final_df.to_excel(writer, sheet_name="最终完整数据", index=False)
                    unmatched_df.to_excel(writer, sheet_name="未匹配数据-计划", index=False)
                    unused_score_df.to_excel(writer, sheet_name="未匹配数据-分数", index=False)

                    ws = writer.book["最终完整数据"]
                    for col_idx, col_name in enumerate(final_df.columns, start=1):
                        if col_name in TEXT_COLUMNS:
                            col_letter = get_column_letter(col_idx)
                            for row in range(2, ws.max_row + 1):
                                ws[f"{col_letter}{row}"].number_format = "@"

                output.seek(0)

                st.download_button(
                    "⬇ 下载 Excel",
                    data=output,
                    file_name=f"匹配结果_{uuid.uuid4().hex[:6]}.xlsx"
                )

                clear_cache()

        except Exception as e:
            st.error(f"❌ 数据处理出错：{str(e)}")
            st.code(traceback.format_exc())
    else:
        st.info("👆 **请上传计划表和分数表开始匹配**")

# =====================================================
# ======================= TAB 5 =======================
# =====================================================
with tab5:
    st.header("📊 学业桥-高考专业分数数据转换")
    st.markdown("""
    ### 📋 功能说明
    本工具用于将"学业桥"系统的专业分数据转换为标准批量导入模板。

    **需要上传以下3个文件：**
    1. **专业分（源数据）** - 从学业桥导出的专业分原始数据
    2. **学校小范围数据导出** - 包含学校名称的标准数据（高考数据库导出）
    3. **专业信息表** - 包含专业名称和层次的数据（高考数据库导出）
    4. **学业桥数据导出时加一列层次放在最后一列**

    5.**处理完的数据需要大体浏览检查一遍**
    """)

    LEVEL_MAP = {
        "1": "本科(普通)",
        "2": "专科(高职)",
        "3": "本科(职业)"
    }

    GROUP_JOIN_PROVINCE = {
        "湖南", "福建", "广东", "北京", "黑龙江", "安徽", "江西", "广西",
        "甘肃", "山西", "河南", "陕西", "宁夏", "四川", "云南", "内蒙古"
    }

    ONLY_CODE_PROVINCE = {
        "湖北", "江苏", "上海", "天津", "海南"
    }

    FINAL_COLUMNS = [
        "学校名称", "省份", "招生专业", "专业方向（选填）", "专业备注（选填）",
        "一级层次", "招生科类", "招生批次", "招生类型（选填）",
        "最高分", "最低分", "平均分",
        "最低分位次（选填）", "招生人数（选填）", "数据来源",
        "专业组代码", "首选科目", "选科要求", "次选科目",
        "专业代码", "招生代码",
        "最低分数区间低", "最低分数区间高",
        "最低分数区间位次低", "最低分数区间位次高",
        "录取人数（选填）"
    ]


    # =========================
    # 工具函数
    # =========================
    def read_file(file):
        name = file.name.lower()

        try:
            if name.endswith(".csv"):
                return pd.read_csv(
                    file,
                    dtype=str,
                    encoding="utf-8",
                    keep_default_na=False  # ✅ 关键
                )

            elif name.endswith(".xlsx"):
                return pd.read_excel(
                    file,
                    dtype=str,
                    engine="openpyxl",
                    keep_default_na=False  # ✅ 关键
                )

            elif name.endswith(".xls"):
                return pd.read_excel(
                    file,
                    dtype=str,
                    engine="xlrd",
                    keep_default_na=False  # ✅ 关键
                )

        except:
            raise ValueError(f"读取失败: {file.name}")

    def to_float(x):
        if pd.isna(x):
            return None
        s = str(x).strip()
        if s == "" or s == "0":
            return None
        try:
            return float(s)
        except:
            return None


    def clean_code(x):
        if pd.isna(x):
            return ""
        x = str(x).strip()
        if x.endswith(".0"):
            x = x[:-2]
        return x



    def score_valid(row):
        max_s = to_float(row["最高分"])
        min_s = to_float(row["最低分"])
        avg_s = to_float(row["平均分"])

        # ❗新增：最低分必须存在且不能为0
        if min_s is None or min_s == 0:
            return False

        # 原有逻辑
        if max_s is not None and max_s < min_s:
            return False

        if avg_s is not None:
            if avg_s < min_s:
                return False
            if max_s is not None and avg_s > max_s:
                return False

        return True


    def convert_subject(x):
        if x == "物理":
            return "物理类", "物"
        if x == "历史":
            return "历史类", "历"
        if x in {"文科", "理科", "综合"}:
            return x, ""
        return x, ""


    def parse_requirement(req):
        if pd.isna(req):
            return "不限科目专业组", ""

        req = str(req).strip()
        if req == "" or req == "不限":
            return "不限科目专业组", ""

        # 次选科目：只保留科目本身
        subjects = req.replace("且", "").replace("/", "")

        # 物/化 → 多门选考
        if "/" in req:
            return "多门选考", subjects

        # 物且化 / 单科
        return "单科、多科均需选考", subjects


    def map_xinjiang_batch(batch, province):
        """
        新疆特殊批次映射
        """
        if province != "新疆":
            return batch

        if pd.isna(batch):
            return batch

        batch_str = str(batch).strip()

        # 映射规则
        if batch_str == "本科一批-其他":
            return "国家及地方专项、南疆单列、对口援疆计划本科一批次"
        elif batch_str == "本科二批-其他":
            return "国家及地方专项、南疆单列、对口援疆计划本科二批次"

        return batch


    def build_group_code(row):
        prov = str(row["省份"]).strip()

        code = clean_code(row["招生代码"])
        gid = clean_code(row["专业组编号"])

        # ❗没有专业组编号 → 直接不给
        if not gid:
            return ""

        # ❗没有招生代码 → 也不给
        if not code:
            return ""

        # ✅ 吉林特殊（无括号）
        if prov == "吉林":
            return f"{code}{gid}"

        # ✅ 需要拼括号的省份（沿用你原来的）
        if prov in GROUP_JOIN_PROVINCE:
            return f"{code}（{gid}）"

        # ✅ 只用招生代码的省份（你原来的规则）
        if prov in ONLY_CODE_PROVINCE:
            return code

        # 默认（不给，避免乱拼）
        return ""


    def to_excel(df):
        buf = BytesIO()
        df.to_excel(buf, index=False)
        return buf.getvalue()


    # =========================
    # 文件上传
    # =========================
    c1, c2, c3 = st.columns(3)
    with c1:
        prof_file = st.file_uploader("📥 上传【专业分（源数据）】", type=["xls", "xlsx", "csv"], key="prof_file_5")
    with c2:
        school_file = st.file_uploader("🏫 学校小范围数据导出", type=["xls", "xlsx"], key="school_file_5")
    with c3:
        major_file = st.file_uploader("📘 专业信息表", type=["xls", "xlsx"], key="major_file_5")

    # =========================
    # 主逻辑
    # =========================
    if prof_file and school_file and major_file:
        df = read_file(prof_file)
        school_df = read_file(school_file)
        major_df = read_file(major_file)

        st.subheader("① 数据校验")

        errors = []

        # 校验1：学校名称
        bad_school = df[~df["院校名称"].isin(set(school_df["学校名称"]))].copy()
        if not bad_school.empty:
            bad_school["错误原因"] = "学校名称不在学校小范围数据中"
            errors.append(bad_school)

        # 校验2：专业 + 一级层次
        df["一级层次"] = df["层次"].map(LEVEL_MAP)
        chk = df.merge(
            major_df[["专业名称", "一级层次"]],
            on=["专业名称", "一级层次"],
            how="left",
            indicator=True
        )
        bad_major = chk[chk["_merge"] == "left_only"].copy()
        if not bad_major.empty:
            bad_major["错误原因"] = "专业名称 + 一级层次 不存在"
            errors.append(bad_major[df.columns.tolist() + ["错误原因"]])

        # 在校验3的错误提示部分修改为：
        bad_score = df[~df.apply(score_valid, axis=1)].copy()
        if not bad_score.empty:
            # 添加更详细的错误原因说明
            def get_score_error(row):
                min_s = to_float(row["最低分"])
                if min_s is None or min_s == 0:
                    return "最低分为空或0"
                # 其他错误...
                return "分数关系错误（最高/平均/最低）"


            bad_score["错误原因"] = bad_score.apply(get_score_error, axis=1)
            errors.append(bad_score)

        if errors:
            err_df = pd.concat(errors, ignore_index=True)
            st.error(f"❌ 校验失败，共 {len(err_df)} 条")
            st.dataframe(err_df)
            st.download_button(
                "📥 下载错误明细",
                data=to_excel(err_df),
                file_name="专业分-校验错误明细.xlsx"
            )
            st.stop()

        st.success("✅ 校验通过")

        # =========================
        # 字段转换
        # =========================
        out = pd.DataFrame()

        out["学校名称"] = df["院校名称"]
        out["省份"] = df["省份"]
        out["招生专业"] = df["专业名称"]
        out["专业方向（选填）"] = ""
        out["专业备注（选填）"] = df["专业备注"]
        out["一级层次"] = df["一级层次"]

        out["招生科类"], out["首选科目"] = zip(*df["科类"].apply(convert_subject))
        out["招生批次"] = df.apply(
            lambda row: map_xinjiang_batch(row["批次"], row["省份"]),
            axis=1
        )

        out["招生类型（选填）"] = df["招生类型"]

        out["最高分"] = df["最高分"]
        out["最低分"] = df["最低分"]
        out["平均分"] = df["平均分"]

        out["最低分位次（选填）"] = ""
        out["招生人数（选填）"] = df["招生计划人数"]

        out["数据来源"] = "学业桥"
        out["专业组代码"] = df.apply(build_group_code, axis=1)

        out["选科要求"], out["次选科目"] = zip(*df["报考要求"].apply(parse_requirement))

        out["专业代码"] = df["专业代码"]
        out["招生代码"] = df["招生代码"]

        out["最低分数区间低"] = ""
        out["最低分数区间高"] = ""
        out["最低分数区间位次低"] = ""
        out["最低分数区间位次高"] = ""

        out["录取人数（选填）"] = df["录取人数"]

        # 强制字段顺序
        out = out[FINAL_COLUMNS]

        st.dataframe(out.head(20))

        st.download_button(
            "📤 下载【专业分-批量导入模板】",
            data=to_excel(out),
            file_name="专业分-批量导入模板.xlsx",
            key="download_result_5"
        )
# ------------------------ 院校分提取处理函数 ------------------------
def process_admission_data(df_source):
    """
    处理院校分数据，按照指定规则分组并生成结果表格
    """
    log("开始处理院校分数据...")

    # 数据清洗和预处理 - 只替换特殊字符，不填充空值
    df_source = df_source.replace({'^': '', '~': ''}, regex=True)

    # 处理数值字段，但不填充空值
    numeric_columns = ['最高分', '最低分', '最低分位次', '录取人数', '招生人数']
    for col in numeric_columns:
        if col in df_source.columns:
            df_source[col] = pd.to_numeric(df_source[col], errors='coerce')

    # 确定首选科目 - 只针对新高考省份
    def determine_preferred_subject(row):
        col_type = str(row.get('科类', ''))
        # 只有历史类和物理类才有首选科目
        if '历史类' in col_type:
            return '历史'
        elif '物理类' in col_type:
            return '物理'
        # 文科、理科、综合等传统科类没有首选科目
        return ''

    df_source['首选科目'] = df_source.apply(determine_preferred_subject, axis=1)

    # 确定招生类别（科类）- 修正逻辑
    def determine_admission_category(row):
        col_type = str(row.get('科类', ''))
        # 新高考省份：历史类、物理类
        if '历史类' in col_type:
            return '历史类'
        elif '物理类' in col_type:
            return '物理类'
        # 传统高考省份：文科、理科
        elif '文科' in col_type:
            return '文科'
        elif '理科' in col_type:
            return '理科'
        elif '综合' in col_type:
            return '综合'
        # 其他情况保持原样
        return col_type

    df_source['招生类别'] = df_source.apply(determine_admission_category, axis=1)

    # 处理层次字段 - 确保不为空
    if '层次' in df_source.columns:
        df_source['层次'] = df_source['层次'].fillna('本科(普通)')
    else:
        df_source['层次'] = '本科(普通)'

    # 处理招生类型 - 确保不为空
    if '招生类型' in df_source.columns:
        df_source['招生类型'] = df_source['招生类型'].fillna('')
    else:
        df_source['招生类型'] = ''

    # 处理专业组代码 - 确保不为空
    if '专业组代码' in df_source.columns:
        df_source['专业组代码'] = (
            df_source['专业组代码']
            .astype(str)
            .str.replace(r'^\^', '', regex=True)  # ⭐ 去掉开头 ^
            .str.strip()
        )
    else:
        df_source['专业组代码'] = ''

    # 处理其他分组列 - 确保不为空
    df_source['省份'] = df_source['省份'].fillna('')
    df_source['批次'] = df_source['批次'].fillna('')
    df_source['学校'] = df_source['学校'].fillna('')

    log("数据预处理完成，开始分组...")

    # 分组处理 - 按照指定的列分组（加上学校）
    grouping_columns = ['学校', '省份', '招生类别', '批次', '层次', '招生类型', '专业组代码']

    log(f"使用以下列进行分组: {grouping_columns}")

    # 创建一个列表来存储结果
    results = []

    # 对每个分组进行处理
    group_count = 0
    for group_key, group_data in df_source.groupby(grouping_columns):
        group_count += 1
        # 解包分组键
        学校, 省份, 招生类别, 批次, 层次, 招生类型, 专业组代码 = group_key

        # 计算组内聚合值 - 根据源数据中是否有该列来决定处理方式
        最高分 = pd.NA
        if '最高分' in group_data.columns and not group_data['最高分'].isna().all():
            最高分 = group_data['最高分'].max()

        最低分 = pd.NA
        if '最低分' in group_data.columns and not group_data['最低分'].isna().all():
            最低分 = group_data['最低分'].min()

        # 找到最低分对应的记录
        最低分位次 = pd.NA
        数据来源 = ''
        首选科目 = ''

        if pd.notna(最低分) and '最低分' in group_data.columns:
            min_score_rows = group_data[group_data['最低分'] == 最低分]
            if not min_score_rows.empty:
                min_score_row = min_score_rows.iloc[0]
                # 这些字段根据源数据决定
                最低分位次 = min_score_row.get('最低分位次', pd.NA) if '最低分位次' in min_score_row else pd.NA
                数据来源 = min_score_row.get('数据来源', '') if '数据来源' in min_score_row else ''
                首选科目 = min_score_row.get('首选科目', '') if '首选科目' in min_score_row else ''

        # 计算录取人数总和（源数据中有录取人数）
        录取人数 = pd.NA
        if '录取人数' in group_data.columns and not group_data['录取人数'].isna().all():
            录取人数 = group_data['录取人数'].sum()

        # 招生人数处理 - 源数据中有就处理，没有就置空
        招生人数 = pd.NA
        if '招生人数' in group_data.columns and not group_data['招生人数'].isna().all():
            招生人数 = group_data['招生人数'].sum()

        # 添加到结果列表 - 只保留指定的列
        result_row = {
            '学校名称': 学校,
            '省份': 省份,
            '招生类别': 招生类别,
            '层次': 层次,
            '招生批次': 批次,
            '招生类型': 招生类型,
            # ⭐ 固定空字段（不参与任何逻辑）
            '选测等级': '',
            '最高分': 最高分,
            '最低分': 最低分,
            '平均分': pd.NA,
            '最高位次': pd.NA,
            '最低位次': 最低分位次,
            '平均位次': pd.NA,
            '最低分位次': 最低分位次,
            '录取人数': 录取人数,
            '招生人数': 招生人数,
            '数据来源': 数据来源,
            # ⭐ 固定空字段
            '省控线科类': '',
            '省控线批次': '',
            '省控线备注': '',
            '专业组代码': 专业组代码,
            '首选科目': 首选科目,
            '院校招生代码': (
                str(
                    min_score_row.get('院校招生代码',
                                      min_score_row.get('招生代码', ''))
                ).replace('^', '').strip()
            ),
            '层次': 层次
        }

        results.append(result_row)

    log(f"分组处理完成，共 {group_count} 个分组")

    # 创建结果DataFrame
    result_df = pd.DataFrame(results)

    log(f"分组后共有 {len(result_df)} 组数据")

    # 确保数值字段保持正确的数据类型
    numeric_columns = ['最高分', '最低分', '最低分位次', '录取人数', '招生人数']
    for col in numeric_columns:
        if col in result_df.columns:
            result_df[col] = pd.to_numeric(result_df[col], errors='coerce')

    log(f"处理完成，共生成 {len(result_df)} 行记录")

    return result_df
# ------------------------tab 6 ------------------------
with tab6:
    st.subheader("院校分提取")
    st.markdown("适用于高考数据库直接导出的专业分数据")
    st.markdown("""
    本工具按照以下规则处理专业分数据：

    - **分组规则**：学校、省份、科类、批次、层次、招生类型、专业组代码

    - **处理规则**：
      - 所有列都根据源数据决定，有值就处理，没值就置空
      - 最高分 = 组内最高分的最大值
      - 最低分 = 组内最低分的最小值
      - 最低分位次 = 最低分对应的位次
      - 录取人数 = 组内录取人数总和
      - 招生人数 = 组内招生人数总和
    """)

    # 文件上传
    uploaded_file_admission = st.file_uploader(
        "上传专业分数据Excel文件",
        type=['xlsx'],
        help="请上传包含专业分数据的Excel文件，系统会输出固定的15列数据",
        key="admission_excel"
    )

    if uploaded_file_admission is not None:
        try:
            # 读取上传的文件
            df_source = pd.read_excel(uploaded_file_admission)

            # 显示源数据信息
            st.subheader("📊 源数据信息")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.write(f"**总记录数:** {len(df_source)}")
            with col2:
                st.write(f"**总列数:** {len(df_source.columns)}")
            with col3:
                st.write(f"**所有列名:** {list(df_source.columns)}")

            # 显示源数据预览
            st.write("**源数据预览:**")
            st.dataframe(df_source.head(10), use_container_width=True)

            # 处理按钮
            if st.button("🚀 开始处理专业分数据", type="primary", key="admission_btn"):
                with st.spinner("正在处理专业分数据，请稍候..."):
                    result_df = process_admission_data(df_source)

                if len(result_df) == 0:
                    st.error("警告：没有生成任何数据，请检查源数据文件")
                    st.stop()

                # 显示处理结果
                st.subheader("✅ 处理结果")

                # 显示统计信息
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("学校数量", result_df['学校名称'].nunique())
                with col2:
                    st.metric("省份数量", result_df['省份'].nunique())
                with col3:
                    st.metric("总记录数", len(result_df))
                with col4:
                    st.metric("输出列数", len(result_df.columns))

                # 显示输出列信息
                st.write(f"**输出列名 ({len(result_df.columns)}列):**")
                output_columns = [
                    '学校名称', '省份', '招生类别', '招生批次', '招生类型', '选测等级',
                    '最高分', '最低分', '平均分',
                    '最高位次', '最低位次', '平均位次',
                    '录取人数', '招生人数', '数据来源',
                    '省控线科类', '省控线批次', '省控线备注',
                    '专业组代码', '首选科目', '院校招生代码', '层次'
                ]
                for i, col in enumerate(output_columns, 1):
                    st.write(f"{i}. {col}")

                # 显示数据预览
                st.dataframe(result_df[output_columns], use_container_width=True)

                # 显示字段统计
                st.subheader("📈 字段数据统计")

                # 检查各字段的有效数据比例
                st.write("**各字段有效数据比例:**")
                stats_data = []
                for col in output_columns:
                    if col in result_df.columns:
                        total = len(result_df)
                        valid = result_df[col].notna().sum()
                        if result_df[col].dtype == 'object':
                            # 对于字符串列，检查非空字符串
                            valid = (result_df[col].notna() & (result_df[col] != '')).sum()
                        percentage = (valid / total) * 100 if total > 0 else 0
                        stats_data.append({
                            '字段名': col,
                            '有效数据数': valid,
                            '有效比例%': f"{percentage:.1f}%"
                        })

                stats_df = pd.DataFrame(stats_data)
                st.dataframe(stats_df, use_container_width=True)

                # 下载功能
                st.subheader("📥 下载处理结果")

                # 将DataFrame转换为Excel字节流，确保列顺序
                output = BytesIO()
                from openpyxl.utils import get_column_letter

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    result_df[output_columns].to_excel(writer, index=False, sheet_name='处理结果')

                    ws = writer.book['处理结果']

                    text_cols = ['专业组代码', '院校招生代码']

                    for col_name in text_cols:
                        if col_name in output_columns:
                            col_idx = output_columns.index(col_name) + 1
                            col_letter = get_column_letter(col_idx)

                            for row in range(2, ws.max_row + 1):
                                ws[f"{col_letter}{row}"].number_format = "@"

                processed_data = output.getvalue()

                st.download_button(
                    label="📥 下载处理后的Excel文件",
                    data=processed_data,
                    file_name="分组处理后的专业分数据.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_admission"
                )

        except Exception as e:
            st.error(f"处理过程中出现错误: {e}")
            st.info("请检查上传的文件格式是否正确")


# =====================================================
# ======================= TAB 7 =======================
# =====================================================
def convert_subject(x):

    if pd.isna(x):
        return "", ""

    x = str(x).strip()

    if x in ["物理", "物理类"]:
        return "物理类", "物"

    if x in ["历史", "历史类"]:
        return "历史类", "历"

    if x in ["综合", "综合改革"]:
        return "综合改革", ""

    if x in ["文科", "理科", "文史", "理工"]:
        return x, ""

    return x, ""


def parse_requirement(req, subject_type):

    subject_type = str(subject_type).strip()

    if subject_type in ["文科", "理科", "文史", "理工"]:
        return "", ""

    if pd.isna(req):
        return "不限科目专业组", ""

    req = str(req).strip()

    if req == "":
        return "不限科目专业组", ""

    if "不限" in req:
        return "不限科目专业组", ""

    req = req.replace(" ", "")

    if "且" in req:
        return (
            "单科、多科均需选考",
            req.replace("且", "")
        )

    seps = [
        "或",
        "/",
        "\\",
        "+",
        "＋",
        "、",
        ",",
        "，"
    ]

    if any(s in req for s in seps):

        tmp = req

        for s in seps:
            tmp = tmp.replace(s, "")

        return "多门选考", tmp

    return "单科、多科均需选考", req



with tab7:
    st.header("📘 学业桥-招生计划模板转换")

    st.markdown("""
    ### 📋 功能说明
    将学业桥导出的数据转换为【招生计划导入模板】

    ### ✅ 强制校验
    - 学校名称校验
    - 专业名称 + 一级层次校验
    - 招生计划人数不能为空或0

    ### ℹ️ 非强制校验（仅备注）
    - 学制校验
    - 学费校验
    """)

    PLAN_COLUMNS = [
        "学校名称",
        "省份",
        "招生专业",
        "专业方向",
        "专业备注",
        "一级层次",
        "招生科类",
        "招生批次",
        "招生类型",
        "招生代码",
        "招生人数",
        "专业学制",
        "学费",
        "数据来源",
        "专业组代码",
        "首选科目",
        "选科要求",
        "次选科目",
        "专业代码",
        "学费单位",
        "批次备注",

        # 新增备注字段
        "学制校验备注",
        "学费校验备注"
    ]

    # =========================
    # 学制校验
    # =========================
    def validate_duration(row):

        major = str(row.get("专业名称", ""))
        remark = str(row.get("专业备注", ""))
        level = str(row.get("一级层次", ""))

        try:
            duration = int(float(row.get("学年", 0)))
        except:
            duration = 0

        # 预科
        if "预科" in major or "预科" in remark:
            return duration == 1

        medical_majors = [
            "临床医学",
            "预防医学",
            "中医学",
            "基础医学",
            "医学影像学",
            "口腔医学",
            "麻醉学",
            "动物医学"
        ]

        is_medical = any(m in major for m in medical_majors)

        # 本科
        if level in ["本科(普通)", "本科(职业)"]:

            if is_medical:
                return duration == 5

            return duration >= 4

        # 专科
        elif level == "专科(高职)":
            return duration == 3

        return True

    # =========================
    # 中外合作识别
    # =========================
    def is_foreign_coop(row):

        foreign_keywords = [
            "中外合作",
            "中日",
            "中英",
            "中澳",
            "中法",
            "中德",
            "中韩",
            "中加",
            "中美"
        ]

        major = str(row.get("专业名称", ""))
        remark = str(row.get("专业备注", ""))

        return any(k in major or k in remark for k in foreign_keywords)

    # =========================
    # 文件上传
    # =========================
    c1, c2, c3 = st.columns(3)

    with c1:
        plan_file = st.file_uploader(
            "📥 上传【学业桥源数据】",
            type=["xls", "xlsx", "csv"],
            key="plan_file_tab7"
        )

    with c2:
        school_file_7 = st.file_uploader(
            "🏫 学校小范围数据导出",
            type=["xls", "xlsx"],
            key="school_file_tab7"
        )

    with c3:
        major_file_7 = st.file_uploader(
            "📘 专业信息表",
            type=["xls", "xlsx"],
            key="major_file_tab7"
        )

    # =========================
    # 主逻辑
    # =========================
    if plan_file and school_file_7 and major_file_7:

        df = read_file(plan_file)

        school_df = read_file(school_file_7)

        major_df = read_file(major_file_7)

        # =========================
        # 层次转换
        # =========================
        df["一级层次"] = df["层次"].map(LEVEL_MAP)

        st.subheader("① 数据校验")

        errors = []

        # =========================
        # 校验1：学校名称
        # =========================
        bad_school = df[
            ~df["院校名称"].isin(set(school_df["学校名称"]))
        ].copy()

        if not bad_school.empty:

            bad_school["错误原因"] = "学校名称不在学校小范围数据中"

            errors.append(bad_school)

        # =========================
        # 校验2：专业 + 一级层次
        # =========================
        chk = df.merge(
            major_df[["专业名称", "一级层次"]],
            on=["专业名称", "一级层次"],
            how="left",
            indicator=True
        )

        bad_major = chk[
            chk["_merge"] == "left_only"
        ].copy()

        if not bad_major.empty:

            bad_major["错误原因"] = "专业名称 + 一级层次 不存在"

            errors.append(
                bad_major[
                    df.columns.tolist() + ["错误原因"]
                ]
            )

        # =========================
        # 校验3：招生计划人数
        # =========================
        def valid_plan_num(x):

            if pd.isna(x):
                return False

            s = str(x).strip()

            if s == "" or s == "0":
                return False

            try:
                return float(s) > 0
            except:
                return False

        bad_plan = df[
            ~df["招生计划人数"].apply(valid_plan_num)
        ].copy()

        if not bad_plan.empty:

            bad_plan["错误原因"] = "招生计划人数为空或0"

            errors.append(bad_plan)

        # =========================
        # 强制校验错误汇总
        # =========================
        if errors:

            err_df = pd.concat(
                errors,
                ignore_index=True
            )

            st.error(f"❌ 校验失败，共 {len(err_df)} 条")

            st.dataframe(err_df)

            st.download_button(
                "📥 下载错误明细",
                data=to_excel(err_df),
                file_name="招生计划-校验错误明细.xlsx"
            )

            st.stop()

        st.success("✅ 强制校验通过")

        # =========================
        # 学制校验（仅备注）
        # =========================
        df["学制校验备注"] = ""

        bad_duration_mask = ~df.apply(
            validate_duration,
            axis=1
        )

        df.loc[
            bad_duration_mask,
            "学制校验备注"
        ] = "专业学制异常"

        # =========================
        # 学费校验（仅备注）
        # =========================

        # 学费数值
        df["学费数值"] = pd.to_numeric(
            df["学费"],
            errors="coerce"
        ).fillna(0)

        # 学费单位
        df["学费单位"] = (
            df["学费单位"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # 中外合作
        df["是否中外合作"] = df.apply(
            is_foreign_coop,
            axis=1
        )

        # 初始化
        df["学费校验"] = True

        df["学费异常原因"] = ""

        # 有效学费
        valid_fee_mask = df["学费数值"] > 0

        # =========================
        # 非中外合作低于2000
        # =========================
        condition1 = (
            (~df["是否中外合作"])
            & valid_fee_mask
            & (df["学费数值"] < 2000)
        )

        df.loc[condition1, "学费校验"] = False

        df.loc[
            condition1,
            "学费异常原因"
        ] += "非中外合作学费低于2000；"

        # =========================
        # 与组均值差超2000
        # =========================
        group_cols = [
            "院校名称",
            "省份",
            "一级层次",
            "是否中外合作"
        ]

        df["组合学费均值"] = (
            df.groupby(group_cols)["学费数值"]
            .transform("mean")
        )

        df["与均值差"] = abs(
            df["学费数值"] - df["组合学费均值"]
        )

        condition2 = (
            valid_fee_mask
            & (df["与均值差"] > 2000)
        )

        df.loc[condition2, "学费校验"] = False

        df.loc[
            condition2,
            "学费异常原因"
        ] += "与组均值差超2000；"

        # =========================
        # 学校内唯一学费
        # =========================
        fee_counts = (
            df[valid_fee_mask]
            .groupby("院校名称")["学费数值"]
            .transform(lambda x: x.map(x.value_counts()))
        )

        df.loc[
            valid_fee_mask,
            "学费出现次数"
        ] = fee_counts

        df["学费出现次数"] = (
            df["学费出现次数"]
            .fillna(0)
        )

        condition3 = (
            (df["学费出现次数"] == 1)
            & valid_fee_mask
        )

        df.loc[condition3, "学费校验"] = False

        df.loc[
            condition3,
            "学费异常原因"
        ] += "学校内学费仅出现1次；"

        # =========================
        # 学费单位异常
        # =========================
        condition4 = (
            valid_fee_mask
            & (
                ~df["学费单位"].isin([
                    "元",
                    "人民币元",
                    "CNY"
                ])
            )
        )

        df.loc[condition4, "学费校验"] = False

        df.loc[
            condition4,
            "学费异常原因"
        ] += "学费单位异常；"

        # =========================
        # 学费备注
        # =========================
        df["学费校验备注"] = ""

        df.loc[
            ~df["学费校验"],
            "学费校验备注"
        ] = df["学费异常原因"]

        # =========================
        # 字段转换
        # =========================
        out = pd.DataFrame()

        out["学校名称"] = df["院校名称"]

        out["省份"] = df["省份"]

        out["招生专业"] = df["专业名称"]

        out["专业方向"] = ""

        out["专业备注"] = df.get("专业备注", "")

        out["一级层次"] = df["一级层次"]

        out["招生科类"], out["首选科目"] = zip(
            *df["科类"].apply(convert_subject)
        )

        out["招生批次"] = df.apply(
            lambda row: map_xinjiang_batch(
                row["批次"],
                row["省份"]
            ),
            axis=1
        )

        out["招生类型"] = df.get("招生类型", "")

        out["招生代码"] = df.get("招生代码", "")

        out["招生人数"] = df["招生计划人数"]

        # 学制
        out["专业学制"] = df.get("学年", "")

        # 学费
        out["学费"] = df.get("学费", "")

        out["数据来源"] = "学业桥"

        out["专业组代码"] = df.apply(
            build_group_code,
            axis=1
        )

        out["选科要求"], out["次选科目"] = zip(
            *df.apply(
                lambda row: parse_requirement(
                    row.get("报考要求", ""),
                    row.get("科类", "")
                ),
                axis=1
            )
        )

        out["专业代码"] = df.get("专业代码", "")

        # 学费单位
        out["学费单位"] = df.get("学费单位", "")

        # 固定空字段
        out["批次备注"] = ""

        # 校验备注
        out["学制校验备注"] = df["学制校验备注"]

        out["学费校验备注"] = df["学费校验备注"]

        # 强制字段顺序
        out = out[PLAN_COLUMNS]

        # =========================
        # 展示结果
        # =========================
        st.subheader("转换结果预览")

        st.dataframe(out.head(20))

        st.download_button(
            "📤 下载【招生计划导入模板】",
            data=to_excel(out),
            file_name="招生计划导入模板.xlsx",
            key="download_plan_tab7"
        )

# ------------------------ Footer ------------------------
st.markdown("---")
st.caption("说明：已默认启用统一请求配置（超时与证书策略）。若需将 VERIFY_SSL 设为 True，请修改文件顶部的常量并重启。")